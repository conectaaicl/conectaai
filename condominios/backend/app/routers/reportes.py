"""
Módulo de Reportes - Exportación Excel y PDF resumen
"""
import io
import os
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.core.database import get_db
from app.core.dependencies import get_current_user

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])

# ── helpers de estilo ──────────────────────────────────────────────────────────

AZUL   = "1E40AF"
CELESTE = "DBEAFE"
VERDE  = "166534"
ROJO   = "991B1B"
GRIS   = "F8FAFC"

def _wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb

def _hdr(ws, row, cols, color=AZUL):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF" if color == AZUL else "1E3A5F", size=11)
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CBD5E1")
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def _row(ws, row, vals, alt=False):
    fill = PatternFill("solid", fgColor=GRIS if alt else "FFFFFF")
    thin = Side(style="thin", color="E2E8F0")
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        c.alignment = Alignment(vertical="center")

def _autofit(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))

def _title(ws, text_val, col_span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    c = ws.cell(row=1, column=1, value=text_val)
    c.font = Font(bold=True, size=14, color="1E40AF")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
    c2 = ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c2.font = Font(size=9, color="64748B")
    c2.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

def _stream(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

# ─────────────────────────────────────────────────────────────────────────────
# 1. MOROSIDAD
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel/morosidad")
def reporte_morosidad(
    condominio_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    tenant_id = current_user["tenant_id"]
    wb = _wb()
    ws = wb.create_sheet("Morosidad")
    ws.sheet_view.showGridLines = True

    _title(ws, "REPORTE DE MOROSIDAD", 8)

    _hdr(ws, 3, ["Torre", "Piso", "Depto", "Residente", "Periodos Adeudados",
                  "Monto Total $", "Ultimo Pago", "Estado"])

    cond_filter = "AND tor.condominio_id=:cid" if condominio_id else ""
    params: dict = {"tid": tenant_id}
    if condominio_id:
        params["cid"] = condominio_id

    rows = db.execute(text(f"""
        SELECT
            COALESCE(tor.nombre, '-') AS torre,
            COALESCE(p.numero::text, '-') AS piso,
            COALESCE(d.numero::text, '-') AS depto,
            COALESCE(gc.nombre_residente, pe.nombre_completo, 'Sin asignar') AS residente,
            COUNT(DISTINCT gc.periodo_id) AS periodos,
            COALESCE(SUM(gc.monto),0) AS total,
            MAX(gc.fecha_pago)::date AS ultimo_pago
        FROM gastos_cobros gc
        LEFT JOIN departamentos d ON d.id = gc.departamento_id
        LEFT JOIN pisos p ON p.id = d.piso_id
        LEFT JOIN torres tor ON tor.id = p.torre_id
        LEFT JOIN personas pe ON pe.id = gc.persona_id
        WHERE gc.tenant_id=:tid
          AND gc.estado IN ('pendiente','vencido')
          {cond_filter}
        GROUP BY tor.nombre, p.numero, d.numero, residente
        HAVING SUM(gc.monto) > 0
        ORDER BY total DESC
    """), params).fetchall()

    total_general = 0
    for i, r in enumerate(rows):
        r = r._mapping
        meses = int(r["periodos"])
        monto = float(r["total"])
        total_general += monto
        if meses >= 3:
            estado = "Critico"
        elif meses >= 2:
            estado = "Alto"
        else:
            estado = "Moderado"
        _row(ws, 4 + i, [
            r["torre"], r["piso"], r["depto"], r["residente"],
            meses, monto,
            str(r["ultimo_pago"]) if r["ultimo_pago"] else "Nunca",
            estado
        ], alt=(i % 2 == 1))

    n = len(rows)
    ws.cell(row=4+n, column=5, value="TOTAL DEUDA:").font = Font(bold=True)
    ws.cell(row=4+n, column=6, value=total_general).font = Font(bold=True, color=ROJO)

    _autofit(ws)
    ws.freeze_panes = "A4"
    return _stream(wb, f"morosidad_{date.today()}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 2. GASTOS COMUNES - PERÍODO
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel/gastos-periodo")
def reporte_gastos_periodo(
    periodo: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    tenant_id = current_user["tenant_id"]
    wb = _wb()

    # Hoja 1: Cobros por departamento
    ws1 = wb.create_sheet("Cobros por Depto")
    _title(ws1, f"GASTOS COMUNES - PERIODO {periodo}", 7)
    _hdr(ws1, 3, ["Depto", "Residente", "Concepto", "Monto $", "Estado",
                   "Fecha Vencimiento", "Fecha Pago"])

    rows = db.execute(text("""
        SELECT gc.depto_numero, gc.nombre_residente, gc.concepto,
               gc.monto, gc.estado, gc.fecha_vencimiento, gc.fecha_pago
        FROM gastos_cobros gc
        JOIN gastos_periodos gp ON gp.id = gc.periodo_id
        WHERE gp.tenant_id=:tid AND gp.periodo=:per
        ORDER BY gc.depto_numero, gc.concepto
    """), {"tid": tenant_id, "per": periodo}).fetchall()

    recaudado = pagado = pendiente = 0
    for i, r in enumerate(rows):
        r = r._mapping
        m = float(r["monto"] or 0)
        if r["estado"] == "pagado":
            pagado += m
        else:
            pendiente += m
        recaudado += m
        _row(ws1, 4+i, [
            r["depto_numero"], r["nombre_residente"], r["concepto"],
            m, r["estado"],
            str(r["fecha_vencimiento"]) if r["fecha_vencimiento"] else "",
            str(r["fecha_pago"])[:10] if r["fecha_pago"] else ""
        ], alt=(i % 2 == 1))

    n = len(rows)
    ws1.cell(row=4+n+1, column=3, value="TOTAL:").font = Font(bold=True)
    ws1.cell(row=4+n+1, column=4, value=recaudado).font = Font(bold=True)
    ws1.cell(row=4+n+2, column=3, value="Pagado:").font = Font(bold=True, color=VERDE)
    ws1.cell(row=4+n+2, column=4, value=pagado).font = Font(bold=True, color=VERDE)
    ws1.cell(row=4+n+3, column=3, value="Pendiente:").font = Font(bold=True, color=ROJO)
    ws1.cell(row=4+n+3, column=4, value=pendiente).font = Font(bold=True, color=ROJO)
    _autofit(ws1)
    ws1.freeze_panes = "A4"

    # Hoja 2: Items (presupuesto)
    ws2 = wb.create_sheet("Items Presupuesto")
    _title(ws2, f"ITEMS - {periodo}", 4)
    _hdr(ws2, 3, ["Concepto", "Categoria", "Monto $", "Notas"])
    items = db.execute(text("""
        SELECT gi.concepto, gi.categoria, gi.monto, gi.notas
        FROM gastos_items gi
        JOIN gastos_periodos gp ON gp.id = gi.periodo_id
        WHERE gp.tenant_id=:tid AND gp.periodo=:per
        ORDER BY gi.categoria, gi.concepto
    """), {"tid": tenant_id, "per": periodo}).fetchall()
    for i, r in enumerate(items):
        r = r._mapping
        _row(ws2, 4+i, [r["concepto"], r["categoria"],
                         float(r["monto"] or 0), r["notas"] or ""], alt=(i%2==1))
    _autofit(ws2)

    return _stream(wb, f"gastos_{periodo}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 3. FINANZAS - INGRESOS VS EGRESOS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel/finanzas")
def reporte_finanzas(
    anio: int = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    tenant_id = current_user["tenant_id"]
    anio = anio or date.today().year
    wb = _wb()
    ws = wb.create_sheet("Finanzas Anual")
    _title(ws, f"RESUMEN FINANCIERO {anio}", 5)
    _hdr(ws, 3, ["Mes", "Ingresos $", "Egresos $", "Balance $", "Morosidad $"])

    meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    tot_ing = tot_egr = tot_bal = tot_mor = 0
    for m in range(1, 13):
        per = f"{anio}-{m:02d}"
        def safe(q, p):
            try:
                r2 = db.execute(text(q), p).fetchone()
                return float(r2[0]) if r2 and r2[0] is not None else 0.0
            except Exception:
                db.rollback()
                return 0.0
        ing = safe("SELECT SUM(gc.monto) FROM gastos_cobros gc JOIN gastos_periodos gp ON gp.id=gc.periodo_id WHERE gp.tenant_id=:tid AND gp.periodo=:per AND gc.estado='pagado'", {"tid": tenant_id, "per": per})
        mor = safe("SELECT SUM(gc.monto) FROM gastos_cobros gc JOIN gastos_periodos gp ON gp.id=gc.periodo_id WHERE gp.tenant_id=:tid AND gp.periodo=:per AND gc.estado IN ('pendiente','vencido')", {"tid": tenant_id, "per": per})
        egr = safe("SELECT SUM(monto) FROM gastos_comun WHERE tenant_id=:tid AND EXTRACT(year FROM fecha)=:yr AND EXTRACT(month FROM fecha)=:mo", {"tid": tenant_id, "yr": anio, "mo": m})
        bal = ing - egr
        tot_ing += ing; tot_egr += egr; tot_bal += bal; tot_mor += mor
        _row(ws, 3+m, [meses[m-1], ing, egr, bal, mor], alt=(m % 2 == 1))
        bal_c = ws.cell(row=3+m, column=4)
        bal_c.font = Font(color=VERDE if bal >= 0 else ROJO, bold=True)

    ws.cell(row=16, column=1, value="TOTAL ANUAL").font = Font(bold=True)
    for ci, v in enumerate([tot_ing, tot_egr, tot_bal, tot_mor], 2):
        c = ws.cell(row=16, column=ci, value=v)
        c.font = Font(bold=True, color=VERDE if ci != 4 or v >= 0 else ROJO)
    _autofit(ws)
    return _stream(wb, f"finanzas_{anio}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 4. PERSONAL / NOMINA
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel/nomina")
def reporte_nomina(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    tenant_id = current_user["tenant_id"]
    wb = _wb()
    ws = wb.create_sheet("Nomina")
    _title(ws, "NOMINA DE PERSONAL", 9)
    _hdr(ws, 3, ["Nombre", "RUT", "Cargo", "Tipo Contrato", "Fecha Ingreso",
                  "Sueldo Base $", "Estado", "Banco", "Cuenta"])

    rows = db.execute(text("""
        SELECT nombre, rut, cargo, tipo_contrato, fecha_ingreso,
               sueldo_base, estado, banco, numero_cuenta
        FROM personal
        WHERE tenant_id=:tid
        ORDER BY nombre
    """), {"tid": tenant_id}).fetchall()

    total_sueldo = 0
    for i, r in enumerate(rows):
        r = r._mapping
        sb = float(r["sueldo_base"] or 0)
        if r["estado"] == "activo":
            total_sueldo += sb
        _row(ws, 4+i, [
            r["nombre"], r["rut"] or "", r["cargo"] or "",
            r["tipo_contrato"] or "", str(r["fecha_ingreso"]) if r["fecha_ingreso"] else "",
            sb, r["estado"], r["banco"] or "", r["numero_cuenta"] or ""
        ], alt=(i%2==1))

    n = len(rows)
    ws.cell(row=4+n+1, column=5, value="TOTAL SUELDO ACTIVOS:").font = Font(bold=True)
    ws.cell(row=4+n+1, column=6, value=total_sueldo).font = Font(bold=True)
    _autofit(ws)
    ws.freeze_panes = "A4"
    return _stream(wb, f"nomina_{date.today()}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 5. INCIDENCIAS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel/incidencias")
def reporte_incidencias(
    desde: Optional[str] = Query(None),
    hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    tenant_id = current_user["tenant_id"]
    if not desde:
        desde = (date.today() - timedelta(days=90)).isoformat()
    if not hasta:
        hasta = date.today().isoformat()

    wb = _wb()
    ws = wb.create_sheet("Incidencias")
    _title(ws, f"REPORTE INCIDENCIAS {desde} al {hasta}", 7)
    _hdr(ws, 3, ["ID", "Titulo", "Tipo", "Prioridad", "Estado",
                  "Reportada", "Resuelta"])

    rows = db.execute(text("""
        SELECT id, titulo, tipo, prioridad, estado,
               created_at::date, fecha_resolucion::date
        FROM incidencias
        WHERE tenant_id=:tid
          AND created_at::date BETWEEN :d AND :h
        ORDER BY created_at DESC
    """), {"tid": tenant_id, "d": desde, "h": hasta}).fetchall()

    for i, r in enumerate(rows):
        r = r._mapping
        _row(ws, 4+i, [
            r["id"], r["titulo"], r["tipo"] or "", r["prioridad"] or "",
            r["estado"],
            str(r["created_at"]) if r["created_at"] else "",
            str(r["fecha_resolucion"]) if r["fecha_resolucion"] else "Pendiente"
        ], alt=(i%2==1))

    _autofit(ws)
    ws.freeze_panes = "A4"
    return _stream(wb, f"incidencias_{date.today()}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 6. RESERVAS
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel/reservas")
def reporte_reservas(
    desde: Optional[str] = Query(None),
    hasta: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    tenant_id = current_user["tenant_id"]
    if not desde:
        desde = date.today().replace(day=1).isoformat()
    if not hasta:
        hasta = date.today().isoformat()

    wb = _wb()
    ws = wb.create_sheet("Reservas")
    _title(ws, f"RESERVAS ESPACIOS COMUNES {desde} al {hasta}", 7)
    _hdr(ws, 3, ["Espacio", "Depto / Residente", "Fecha", "Hora Inicio", "Hora Fin",
                  "Estado", "Cobro $"])

    rows = db.execute(text("""
        SELECT ec.nombre AS espacio, r.nombre_solicitante, r.departamento_num,
               r.fecha, r.hora_inicio, r.hora_fin, r.estado, r.costo
        FROM reservas r
        JOIN espacios_comun ec ON ec.id = r.espacio_id
        WHERE r.tenant_id=:tid AND r.fecha BETWEEN :d AND :h
        ORDER BY r.fecha, r.hora_inicio
    """), {"tid": tenant_id, "d": desde, "h": hasta}).fetchall()

    total_cobro = 0
    for i, r in enumerate(rows):
        r = r._mapping
        cobro = float(r["costo"] or 0)
        total_cobro += cobro
        _row(ws, 4+i, [
            r["espacio"],
            f"{r['departamento_num']} - {r['nombre_solicitante']}",
            str(r["fecha"]),
            str(r["hora_inicio"]) if r["hora_inicio"] else "",
            str(r["hora_fin"]) if r["hora_fin"] else "",
            r["estado"], cobro
        ], alt=(i%2==1))

    n = len(rows)
    ws.cell(row=4+n+1, column=6, value="TOTAL COBROS:").font = Font(bold=True)
    ws.cell(row=4+n+1, column=7, value=total_cobro).font = Font(bold=True)
    _autofit(ws)
    ws.freeze_panes = "A4"
    return _stream(wb, f"reservas_{date.today()}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 7. RESIDENTES
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/excel/residentes")
def reporte_residentes(
    condominio_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    tenant_id = current_user["tenant_id"]
    wb = _wb()
    ws = wb.create_sheet("Residentes")
    _title(ws, "PADRON DE RESIDENTES Y PROPIETARIOS", 8)
    _hdr(ws, 3, ["Nombre", "RUT", "Tipo", "Depto", "Torre", "Email", "Telefono", "Estado"])

    cond_filter = "AND (datos_contacto->>'condominio_id')::int=:cid" if condominio_id else ""
    params: dict = {"tid": tenant_id}
    if condominio_id:
        params["cid"] = condominio_id

    rows = db.execute(text(f"""
        SELECT nombre_completo, rut, tipo,
               datos_contacto->>'departamento' AS depto,
               datos_contacto->>'torre' AS torre,
               email, telefono, estado
        FROM personas
        WHERE tenant_id=:tid {cond_filter}
        ORDER BY nombre_completo
    """), params).fetchall()

    for i, r in enumerate(rows):
        r = r._mapping
        _row(ws, 4+i, [
            r["nombre_completo"], r["rut"] or "", r["tipo"] or "",
            r["depto"] or "", r["torre"] or "",
            r["email"] or "", r["telefono"] or "", r["estado"]
        ], alt=(i%2==1))

    _autofit(ws)
    ws.freeze_panes = "A4"
    return _stream(wb, f"residentes_{date.today()}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 8. STATS GENERALES (JSON para graficos)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/stats/general")
def stats_general(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    tenant_id = current_user["tenant_id"]
    hoy = date.today()
    per_actual = hoy.strftime("%Y-%m")

    def safe(q, p):
        try:
            r = db.execute(text(q), p).fetchone()
            return r[0] if r and r[0] is not None else 0
        except Exception:
            db.rollback()
            return 0

    total_deptos = safe(
        "SELECT COUNT(*) FROM departamentos d JOIN pisos p ON p.id=d.piso_id JOIN torres t ON t.id=p.torre_id WHERE t.tenant_id=:tid",
        {"tid": tenant_id})
    total_personas = safe(
        "SELECT COUNT(*) FROM personas WHERE tenant_id=:tid AND estado='activo'",
        {"tid": tenant_id})
    deuda_total = safe(
        "SELECT SUM(monto) FROM gastos_cobros gc JOIN gastos_periodos gp ON gp.id=gc.periodo_id WHERE gp.tenant_id=:tid AND gc.estado IN ('pendiente','vencido')",
        {"tid": tenant_id})
    recaudado_mes = safe(
        "SELECT SUM(gc.monto) FROM gastos_cobros gc JOIN gastos_periodos gp ON gp.id=gc.periodo_id WHERE gp.tenant_id=:tid AND gp.periodo=:per AND gc.estado='pagado'",
        {"tid": tenant_id, "per": per_actual})
    incidencias_abiertas = safe(
        "SELECT COUNT(*) FROM incidencias WHERE tenant_id=:tid AND estado NOT IN ('resuelta','cerrada')",
        {"tid": tenant_id})
    visitas_hoy = safe(
        "SELECT COUNT(*) FROM visitas WHERE tenant_id=:tid AND DATE(fecha_ingreso)=:hoy",
        {"tid": tenant_id, "hoy": hoy})
    personal_activo = safe(
        "SELECT COUNT(*) FROM personal WHERE tenant_id=:tid AND estado='activo'",
        {"tid": tenant_id})
    reservas_mes = safe(
        "SELECT COUNT(*) FROM reservas WHERE tenant_id=:tid AND EXTRACT(year FROM fecha)=:yr AND EXTRACT(month FROM fecha)=:mo",
        {"tid": tenant_id, "yr": hoy.year, "mo": hoy.month})

    # Trend morosidad vs recaudacion ultimos 6 meses
    trend = []
    for i in range(5, -1, -1):
        yr = hoy.year
        mo = hoy.month - i
        while mo <= 0:
            mo += 12
            yr -= 1
        p = f"{yr}-{mo:02d}"
        mor = float(safe(
            "SELECT SUM(gc.monto) FROM gastos_cobros gc JOIN gastos_periodos gp ON gp.id=gc.periodo_id WHERE gp.tenant_id=:tid AND gp.periodo=:per AND gc.estado IN ('pendiente','vencido')",
            {"tid": tenant_id, "per": p}) or 0)
        rec = float(safe(
            "SELECT SUM(gc.monto) FROM gastos_cobros gc JOIN gastos_periodos gp ON gp.id=gc.periodo_id WHERE gp.tenant_id=:tid AND gp.periodo=:per AND gc.estado='pagado'",
            {"tid": tenant_id, "per": p}) or 0)
        trend.append({"periodo": p, "morosidad": mor, "recaudado": rec})

    return {
        "totales": {
            "departamentos": int(total_deptos),
            "residentes_activos": int(total_personas),
            "deuda_total": float(deuda_total),
            "recaudado_mes": float(recaudado_mes),
            "incidencias_abiertas": int(incidencias_abiertas),
            "visitas_hoy": int(visitas_hoy),
            "personal_activo": int(personal_activo),
            "reservas_mes": int(reservas_mes),
        },
        "trend_6_meses": trend
    }
