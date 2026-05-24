"""
Motor de Migracion — ConectaAI Condominios
POST /api/migracion/analizar-excel  → analizar Excel/CSV, detectar columnas
POST /api/migracion/validar         → validar datos mapeados
POST /api/migracion/importar        → ejecutar importacion
POST /api/migracion/analizar-pdf    → OCR + IA extraction de PDF
GET  /api/migracion/historial       → historial de importaciones
GET  /api/migracion/plantilla/{tipo}→ descargar plantilla Excel
POST /api/migracion/ai-mapear       → sugerir mapeo con Claude
"""
import io, re, json, os, tempfile
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional, List, Dict, Any
from pydantic import BaseModel
from app.core.database import get_db

router = APIRouter(prefix="/api/migracion", tags=["Migracion"])

TIPOS_VALIDOS = ["residentes", "vehiculos", "gastos_comunes"]

CAMPOS_POR_TIPO = {
    "residentes": ["nombre_completo", "depto_numero", "rut", "email", "telefono", "tipo", "ignorar"],
    "vehiculos": ["patente", "depto_numero", "marca", "modelo", "color", "ignorar"],
    "gastos_comunes": ["depto_numero", "monto", "periodo", "estado", "ignorar"],
}

SUGERENCIAS_AUTO = {
    "nombre": "nombre_completo", "name": "nombre_completo", "propietario": "nombre_completo",
    "residente": "nombre_completo", "arrendatario": "nombre_completo",
    "depto": "depto_numero", "departamento": "depto_numero", "unidad": "depto_numero",
    "apto": "depto_numero", "apt": "depto_numero", "dpto": "depto_numero", "numero": "depto_numero",
    "rut": "rut", "dni": "rut", "ci": "rut", "cedula": "rut", "documento": "rut",
    "email": "email", "correo": "email", "mail": "email",
    "tel": "telefono", "telefono": "telefono", "phone": "telefono",
    "celular": "telefono", "movil": "telefono", "fono": "telefono",
    "tipo": "tipo", "categoria": "tipo", "clasificacion": "tipo",
    "piso": "piso", "floor": "piso",
    "area": "area_m2", "m2": "area_m2", "superficie": "area_m2",
    "patente": "patente", "placa": "patente", "matricula": "patente",
    "marca": "marca", "marca_vehiculo": "marca",
    "modelo": "modelo", "modelo_vehiculo": "modelo",
    "color": "color", "color_vehiculo": "color",
    "monto": "monto", "valor": "monto", "importe": "monto", "deuda": "monto",
    "periodo": "periodo", "mes": "periodo", "fecha": "periodo",
    "estado": "estado", "status": "estado",
}


def _ensure_tables(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS migracion_historial (
            id SERIAL PRIMARY KEY,
            tenant_id INTEGER NOT NULL,
            tipo VARCHAR(50) NOT NULL,
            archivo_nombre VARCHAR(300),
            total_filas INTEGER DEFAULT 0,
            importados INTEGER DEFAULT 0,
            errores INTEGER DEFAULT 0,
            estado VARCHAR(20) DEFAULT 'completado',
            usuario VARCHAR(200),
            detalle_errores JSONB,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
    """))
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS vehiculos (
            id SERIAL PRIMARY KEY,
            tenant_id INTEGER NOT NULL,
            depto_numero VARCHAR(20),
            patente VARCHAR(20) NOT NULL,
            marca VARCHAR(100),
            modelo VARCHAR(100),
            color VARCHAR(100),
            activo BOOLEAN DEFAULT true,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
    """))
    db.commit()


def _sugerir_mapeo(columnas: list[str], tipo: str) -> dict:
    mapeo = {}
    campos_disponibles = CAMPOS_POR_TIPO.get(tipo, [])
    for col in columnas:
        col_lower = col.lower().strip().replace(" ", "_").replace("-", "_")
        sugerido = SUGERENCIAS_AUTO.get(col_lower)
        if sugerido and sugerido in campos_disponibles:
            mapeo[col] = sugerido
        else:
            for key, val in SUGERENCIAS_AUTO.items():
                if key in col_lower and val in campos_disponibles:
                    mapeo[col] = val
                    break
    return mapeo


def _validar_email(email: str) -> bool:
    if not email:
        return True
    return bool(re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', str(email).strip()))


def _validar_rut(rut: str) -> bool:
    if not rut:
        return True
    rut = str(rut).upper().replace(".", "").replace("-", "").strip()
    if len(rut) < 2:
        return False
    cuerpo, dv = rut[:-1], rut[-1]
    if not cuerpo.isdigit():
        return False
    total, factor = 0, 2
    for d in reversed(cuerpo):
        total += int(d) * factor
        factor = 9 if factor == 7 else factor + 1
    resto = total % 11
    dv_calc = "K" if resto == 1 else "0" if resto == 0 else str(11 - resto)
    return dv == dv_calc


def _validar_patente(patente: str) -> bool:
    if not patente:
        return False
    p = str(patente).upper().strip().replace(" ", "").replace("-", "")
    return bool(re.match(r'^[A-Z]{2}\d{4}$', p) or re.match(r'^[A-Z]{4}\d{2}$', p) or re.match(r'^[A-Z]{3}\d{3}$', p))


def _validar_telefono(tel: str) -> bool:
    if not tel:
        return True
    tel_clean = re.sub(r'[\s\-\+\(\)]', '', str(tel))
    return bool(re.match(r'^(\+?56)?9?\d{8}$', tel_clean))


def _validar_fila(fila: dict, tipo: str, mapeo: dict) -> tuple[list, list]:
    errores, advertencias = [], []
    datos_mapeados = {}
    for col_orig, campo in mapeo.items():
        if campo != "ignorar":
            datos_mapeados[campo] = fila.get(col_orig, "")

    if tipo == "residentes":
        if not datos_mapeados.get("nombre_completo"):
            errores.append("Nombre completo es obligatorio")
        if not datos_mapeados.get("depto_numero"):
            errores.append("Numero de departamento es obligatorio")
        email = str(datos_mapeados.get("email", "") or "")
        if email and not _validar_email(email):
            errores.append(f"Email invalido: {email}")
        rut = str(datos_mapeados.get("rut", "") or "")
        if rut and not _validar_rut(rut):
            advertencias.append(f"RUT con formato no estandar: {rut}")
        tel = str(datos_mapeados.get("telefono", "") or "")
        if tel and not _validar_telefono(tel):
            advertencias.append(f"Telefono con formato inusual: {tel}")

    elif tipo == "vehiculos":
        patente = str(datos_mapeados.get("patente", "") or "")
        if not patente:
            errores.append("Patente es obligatoria")
        elif not _validar_patente(patente):
            advertencias.append(f"Formato de patente inusual: {patente}")
        if not datos_mapeados.get("depto_numero"):
            errores.append("Numero de departamento es obligatorio")

    elif tipo == "gastos_comunes":
        if not datos_mapeados.get("depto_numero"):
            errores.append("Numero de departamento es obligatorio")
        monto = datos_mapeados.get("monto", "")
        if monto:
            try:
                float(str(monto).replace(",", ".").replace("$", "").strip())
            except ValueError:
                errores.append(f"Monto no es un numero: {monto}")
        else:
            errores.append("Monto es obligatorio")

    return errores, advertencias


# ─── ENDPOINTS ──────────────────────────────────────────────────────────────

@router.post("/analizar-excel", status_code=200)
async def analizar_excel(
    file: UploadFile = File(...),
    tenant_id: int = Form(...),
    tipo: str = Form("residentes"),
    db: Session = Depends(get_db)
):
    _ensure_tables(db)
    content = await file.read()
    filename = file.filename or "archivo.xlsx"
    ext = filename.rsplit(".", 1)[-1].lower()

    try:
        import pandas as pd
        if ext == "csv":
            for enc in ["utf-8", "latin-1", "cp1252"]:
                try:
                    df = pd.read_csv(io.BytesIO(content), encoding=enc, dtype=str)
                    break
                except Exception:
                    continue
            else:
                raise HTTPException(400, "No se pudo leer el CSV")
        else:
            df = pd.read_excel(io.BytesIO(content), dtype=str)

        df = df.fillna("").rename(columns=lambda c: str(c).strip())
        columnas = list(df.columns)
        preview = df.head(5).to_dict(orient="records")
        datos = df.to_dict(orient="records")
        return {
            "columnas": columnas,
            "total_filas": len(df),
            "preview": preview,
            "datos": datos,
            "sugerencias_mapeo": _sugerir_mapeo(columnas, tipo),
            "archivo_nombre": filename,
        }
    except ImportError:
        raise HTTPException(503, "pandas no instalado. Instale: pip install pandas openpyxl")
    except Exception as e:
        raise HTTPException(400, f"Error al leer archivo: {str(e)}")


class ValidarRequest(BaseModel):
    tenant_id: int
    tipo: str
    mapeo: Dict[str, str]
    datos: List[Dict[str, Any]]


@router.post("/validar")
def validar_datos(body: ValidarRequest, db: Session = Depends(get_db)):
    _ensure_tables(db)
    if body.tipo not in TIPOS_VALIDOS:
        raise HTTPException(400, "Tipo invalido: " + body.tipo)

    filas_resultado = []
    validos = con_advertencias = errores_criticos = 0

    emails_vistos, ruts_vistos, deptos_nombres = set(), set(), set()

    for i, fila in enumerate(body.datos):
        errores, advertencias = _validar_fila(fila, body.tipo, body.mapeo)

        datos_mapeados = {v: fila.get(k, "") for k, v in body.mapeo.items() if v != "ignorar"}

        if body.tipo == "residentes":
            email = str(datos_mapeados.get("email", "") or "").lower().strip()
            if email and email in emails_vistos:
                advertencias.append(f"Email duplicado en esta importacion: {email}")
            elif email:
                emails_vistos.add(email)

            rut = str(datos_mapeados.get("rut", "") or "").upper().replace(".", "").replace("-", "").strip()
            if rut and rut in ruts_vistos:
                errores.append(f"RUT duplicado en esta importacion: {rut}")
            elif rut:
                ruts_vistos.add(rut)

            clave = str(datos_mapeados.get("nombre_completo", "")).lower() + "|" + str(datos_mapeados.get("depto_numero", ""))
            if clave in deptos_nombres:
                advertencias.append("Posible duplicado: mismo nombre y depto")
            else:
                deptos_nombres.add(clave)

        if errores:
            estado = "error"
            errores_criticos += 1
        elif advertencias:
            estado = "advertencia"
            con_advertencias += 1
        else:
            estado = "ok"
            validos += 1

        filas_resultado.append({
            "fila_num": i + 2,
            "datos": datos_mapeados,
            "errores": errores,
            "advertencias": advertencias,
            "estado": estado,
        })

    return {
        "validos": validos,
        "con_advertencias": con_advertencias,
        "errores_criticos": errores_criticos,
        "total": len(body.datos),
        "filas": filas_resultado,
    }


class ImportarRequest(BaseModel):
    tenant_id: int
    tipo: str
    filas: List[Dict[str, Any]]
    usuario: str = "admin"


@router.post("/importar")
def importar_datos(body: ImportarRequest, db: Session = Depends(get_db)):
    _ensure_tables(db)
    if body.tipo not in TIPOS_VALIDOS:
        raise HTTPException(400, "Tipo invalido")

    importados, errores_count = 0, 0
    ids_creados, detalles_errores = [], []

    for fila in body.filas:
        datos = fila.get("datos", fila)
        try:
            if body.tipo == "residentes":
                datos_contacto = json.dumps({
                    "email": str(datos.get("email", "") or ""),
                    "telefono": str(datos.get("telefono", "") or ""),
                    "departamento": str(datos.get("depto_numero", "") or ""),
                })
                row = db.execute(text(
                    "INSERT INTO personas (tenant_id, nombre_completo, rut, tipo, estado, datos_contacto) "
                    "VALUES (:tid, :nom, :rut, :tipo, 'activo', CAST(:dc AS jsonb)) "
                    "ON CONFLICT DO NOTHING RETURNING id"
                ), {
                    "tid": body.tenant_id,
                    "nom": str(datos.get("nombre_completo", "") or ""),
                    "rut": str(datos.get("rut", "") or "") or None,
                    "tipo": str(datos.get("tipo", "residente") or "residente"),
                    "dc": datos_contacto,
                }).fetchone()
                if row:
                    ids_creados.append(row._mapping["id"])
                    importados += 1

            elif body.tipo == "vehiculos":
                pat = str(datos.get("patente", "") or "").upper().replace(" ", "").replace("-", "")
                row = db.execute(text(
                    "INSERT INTO vehiculos (tenant_id, depto_numero, patente, marca, modelo, color) "
                    "VALUES (:tid, :dep, :pat, :mar, :mod, :col) "
                    "ON CONFLICT DO NOTHING RETURNING id"
                ), {
                    "tid": body.tenant_id, "dep": str(datos.get("depto_numero", "") or ""),
                    "pat": pat, "mar": str(datos.get("marca", "") or ""),
                    "mod": str(datos.get("modelo", "") or ""), "col": str(datos.get("color", "") or ""),
                }).fetchone()
                if row:
                    ids_creados.append(row._mapping["id"])
                    importados += 1

            elif body.tipo == "gastos_comunes":
                monto_str = str(datos.get("monto", "0") or "0").replace(",", ".").replace("$", "").strip()
                monto = float(monto_str) if monto_str else 0
                db.execute(text(
                    "INSERT INTO migracion_gastos_comunes (tenant_id, depto_numero, monto, periodo, estado) "
                    "VALUES (:tid, :dep, :monto, :periodo, :est) ON CONFLICT DO NOTHING"
                ), {
                    "tid": body.tenant_id, "dep": str(datos.get("depto_numero", "") or ""),
                    "monto": monto, "periodo": str(datos.get("periodo", "") or ""),
                    "est": str(datos.get("estado", "pendiente") or "pendiente"),
                })
                importados += 1

        except Exception as e:
            db.rollback()
            errores_count += 1
            detalles_errores.append({"datos": datos, "error": str(e)[:200]})
            continue

    db.execute(text(
        "INSERT INTO migracion_historial (tenant_id, tipo, total_filas, importados, errores, usuario) "
        "VALUES (:tid, :tipo, :tot, :imp, :err, :usr)"
    ), {
        "tid": body.tenant_id, "tipo": body.tipo,
        "tot": len(body.filas), "imp": importados,
        "err": errores_count, "usr": body.usuario,
    })
    db.commit()
    return {"importados": importados, "errores": errores_count, "ids_creados": ids_creados, "detalles_errores": detalles_errores}


@router.post("/analizar-pdf")
async def analizar_pdf(
    file: UploadFile = File(...),
    tenant_id: int = Form(...),
    db: Session = Depends(get_db)
):
    content = await file.read()
    filename = file.filename or "doc.pdf"
    ext = filename.rsplit(".", 1)[-1].lower()
    texto_ocr = ""

    try:
        if ext == "pdf":
            try:
                from pdf2image import convert_from_bytes
                import pytesseract
                images = convert_from_bytes(content, dpi=200)
                for img in images[:4]:
                    texto_ocr += pytesseract.image_to_string(img, lang="spa") + "\n"
            except ImportError:
                try:
                    import pdfplumber
                    import io as _io
                    with pdfplumber.open(_io.BytesIO(content)) as pdf:
                        for page in pdf.pages[:4]:
                            texto_ocr += (page.extract_text() or "") + "\n"
                except ImportError:
                    raise HTTPException(503, "Instale pdf2image+pytesseract o pdfplumber para procesar PDFs")
        else:
            try:
                import pytesseract
                from PIL import Image
                img = Image.open(io.BytesIO(content))
                texto_ocr = pytesseract.image_to_string(img, lang="spa")
            except ImportError:
                raise HTTPException(503, "Instale pytesseract para procesar imagenes")

        datos_detectados = {"residentes": [], "vehiculos": []}
        confianza = "baja"

        if texto_ocr.strip():
            try:
                import anthropic
                client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
                prompt = (
                    "Eres un sistema de extraccion de datos de condominios. Del siguiente texto extraido de un documento, "
                    "identifica y extrae en JSON estricto:\n"
                    "- residentes: lista de {nombre, depto, rut, email, telefono, tipo}\n"
                    "- vehiculos: lista de {patente, marca, modelo, depto}\n"
                    "Usa string vacio para campos no encontrados. Responde SOLO con JSON valido, sin markdown ni explicacion.\n\n"
                    "TEXTO:\n" + texto_ocr[:3000]
                )
                resp = client.messages.create(
                    model="claude-haiku-4-5-20251001",
                    max_tokens=2000,
                    messages=[{"role": "user", "content": prompt}]
                )
                raw = resp.content[0].text.strip()
                if raw.startswith("```"):
                    raw = raw.split("```")[1]
                    if raw.startswith("json"):
                        raw = raw[4:]
                datos_detectados = json.loads(raw)
                total = len(datos_detectados.get("residentes", [])) + len(datos_detectados.get("vehiculos", []))
                confianza = "alta" if total > 5 else "media" if total > 0 else "baja"
            except Exception:
                confianza = "media" if texto_ocr.strip() else "baja"

        return {
            "texto_ocr": texto_ocr[:500],
            "datos_detectados": datos_detectados,
            "confianza": confianza,
            "total_detectados": sum(len(v) for v in datos_detectados.values() if isinstance(v, list)),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error al procesar archivo: {str(e)}")


@router.get("/historial")
def historial(tenant_id: int, limit: int = 20, db: Session = Depends(get_db)):
    _ensure_tables(db)
    rows = db.execute(text(
        "SELECT * FROM migracion_historial WHERE tenant_id=:tid ORDER BY created_at DESC LIMIT :lim"
    ), {"tid": tenant_id, "lim": limit}).fetchall()
    result = []
    for r in rows:
        d = dict(r._mapping)
        d["created_at"] = str(d.get("created_at") or "")
        if d.get("detalle_errores") and isinstance(d["detalle_errores"], str):
            try:
                d["detalle_errores"] = json.loads(d["detalle_errores"])
            except Exception:
                pass
        result.append(d)
    return result


@router.get("/plantilla/{tipo}")
def descargar_plantilla(tipo: str):
    if tipo not in TIPOS_VALIDOS:
        raise HTTPException(400, "Tipo invalido")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active

        PLANTILLAS = {
            "residentes": {
                "headers": ["Nombre Completo", "Depto/Unidad", "RUT", "Email", "Telefono", "Tipo (propietario/arrendatario)"],
                "ejemplos": [
                    ["Juan Perez Rodriguez", "101", "12.345.678-9", "juan@email.com", "+56912345678", "propietario"],
                    ["Maria Gonzalez", "202", "9.876.543-2", "maria@email.com", "+56987654321", "arrendatario"],
                ],
            },
            "vehiculos": {
                "headers": ["Patente", "Depto/Unidad", "Marca", "Modelo", "Color"],
                "ejemplos": [
                    ["ABCD12", "101", "Toyota", "Corolla", "Blanco"],
                    ["XY1234", "202", "Hyundai", "Accent", "Gris"],
                ],
            },
            "gastos_comunes": {
                "headers": ["Depto/Unidad", "Monto", "Periodo (YYYY-MM)", "Estado"],
                "ejemplos": [
                    ["101", "85000", "2024-01", "pendiente"],
                    ["202", "92000", "2024-01", "pagado"],
                ],
            },
        }

        datos = PLANTILLAS[tipo]
        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(bold=True, color="FFFFFF")

        for col_idx, header in enumerate(datos["headers"], 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64 + col_idx)].width = max(len(header) + 4, 18)

        for row_idx, ejemplo in enumerate(datos["ejemplos"], 2):
            for col_idx, val in enumerate(ejemplo, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=plantilla_{tipo}.xlsx"},
        )
    except ImportError:
        raise HTTPException(503, "openpyxl no instalado")


class AiMapearRequest(BaseModel):
    columnas: List[str]
    tipo: str


@router.post("/ai-mapear")
def ai_mapear(body: AiMapearRequest):
    mapeo_base = _sugerir_mapeo(body.columnas, body.tipo)
    campos = CAMPOS_POR_TIPO.get(body.tipo, [])
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        prompt = (
            f"Tienes columnas de un Excel: {body.columnas}\n"
            f"Para importar '{body.tipo}', los campos disponibles son: {campos}\n"
            f"Sugiere el mapeo en JSON dict donde key=columna_original, value=campo_destino (o 'ignorar').\n"
            f"Responde SOLO con JSON valido."
        )
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = resp.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        mapeo_ia = json.loads(raw)
        mapeo_base.update(mapeo_ia)
    except Exception:
        pass
    return mapeo_base
